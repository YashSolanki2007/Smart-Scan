# Imports
from flask import Flask, render_template, request, url_for
import requests
import json
from tensorflow.keras.preprocessing.text import Tokenizer
from tensorflow.keras.preprocessing.sequence import pad_sequences
from openpyxl import Workbook
from PIL import ImageGrab
import PIL
import tensorflow as tf
from tensorflow import keras
import numpy as np
from tensorflow.keras.preprocessing import image
from tensorflow.keras.models import load_model


# Creating the app
app = Flask(__name__)

# Global Variables
model_prediction = ""
total_cases = ""
total_deaths = ""
country_total_cases = ""
country_total_deaths = ""
country_new_cases = ""
country_new_deaths = ""
country_name_in_response = ""
news_1_title = ""
news_1_link = ""
news_2_title = ""
news_2_link = ""
news_3_title = ""
news_3_link = ""
news_4_title = ""
news_4_link = ""
news_5_title = ""
news_5_link = ""
msg = ""
reply = ""

# Making the home page
@app.route("/")
def home():
    return render_template("home.html")

# Making the prediction of covid or not page
@app.route("/covid-predictor", methods=['GET', 'POST'])
def covid_predictor():
    global model_prediction

    if request.method == "POST":
        # Taking the snapshot
        snapshot = PIL.ImageGrab.grab(bbox=(735, 324, 2145, 1325))
        save_path = "test" + ".png"
        snapshot.save(save_path)

        # Making the list of class names
        class_names = ["Covid", "Normal"]

        # Loading the saved model
        model = load_model("covid-model-2/")

        # Compiling the model
        model.compile(optimizer='adam', loss='sparse_categorical_crossentropy', metrics=['accuracy'])


        test_img = "test.png"
        img = image.load_img(test_img, target_size=(300, 300))

        X = image.img_to_array(img)
        X = np.expand_dims(X, axis=[0])
        images = np.vstack([X])

        prediction = model.predict(images)
        model_prediction = class_names[np.argmax(prediction)]

    return render_template("covid-predictor.html", prediction=model_prediction)

# Making the covid case tracker page
@app.route("/covid-case-checker", methods=['GET', 'POST'])
def covid_case_counter():
    global total_cases, total_deaths, country_total_cases, country_total_deaths, country_new_cases, country_new_deaths, country_name_in_response
    if request.method == "POST":
        webpage = request.form
        country_name = str(webpage['countryInp'])
        print(country_name)
        # Specify the url for the api
        url = "https://api.covid19api.com/summary"

        # Fetch the api data
        response = requests.get(url)

        # Taking the response and converting it into JSON format
        jsonated_response = response.json()

        countries_list = jsonated_response['Countries']

        # Loop over the countries
        for i in range(len(countries_list)):
            country_name_in_res = countries_list[i]['Country']
            country_new_cases = countries_list[i]['NewConfirmed']
            country_new_deaths = countries_list[i]['NewDeaths']
            country_total_cases = countries_list[i]['TotalConfirmed']
            country_total_deaths = countries_list[i]['TotalDeaths']

            if country_name == country_name_in_res:
                print(
                    f"The name of the country is: {country_name_in_res} the new confiremd cases are: {country_new_cases} and the new deaths are: {country_new_deaths} and total confirmed cases are {country_total_cases} also the toal number of deaths are: {country_total_deaths}")
                break

            else:
                pass

        # print(jsonated_response)

    return render_template("covid-case-counter.html", new_cases=country_new_cases, new_deaths=country_new_deaths,
                           total_cases=country_total_cases, total_deaths=country_total_deaths)


# Current Covid News
@app.route("/vaccine-news-page", methods=['GET', 'POST'])
def get_covid_vaccine_news():

    global news_1_title, news_1_link, news_2_title, news_2_link, news_3_title, news_3_link, news_4_title, news_4_link, news_5_title, news_5_link

    url = "https://vaccovid-coronavirus-vaccine-and-treatment-tracker.p.rapidapi.com/api/news/get-vaccine-news/0"

    headers = {
        'x-rapidapi-host': "vaccovid-coronavirus-vaccine-and-treatment-tracker.p.rapidapi.com",
        'x-rapidapi-key': "77ee9999bdmsh61f74ea752bc214p12f26cjsn3bc1f1725b22"
    }

    response = requests.request("GET", url, headers=headers)

    jsonated_response = response.json()

    news = jsonated_response['news']

    news_1_title = news[0]['title']
    news_1_link = news[0]['link']
    news_2_title = news[1]['title']
    news_2_link = news[1]['link']
    news_3_title = news[2]['title']
    news_3_link = news[2]['link']
    news_4_title = news[3]['title']
    news_4_link = news[3]['link']
    news_5_title = news[4]['title']
    news_5_link = news[4]['link']

    if request.method == "POST":
        wb = Workbook()
        ws = wb.active
        for i in range(1, len(news)):
            # row = f"A{i}"
            # print(row)
            ws[f'A{i}'] = "Title: " + news[i]['title']
            ws[f'B{i}'] = news[i]['link']
        # Save the file
        wb.save("sample.xlsx")

    return render_template("vaccine-news-page.html", news_1=news_1_title, news_1_link=news_1_link, news_2=news_2_title, news_2_link=news_2_link, news_3=news_3_title, news_3_link=news_3_link, news_4=news_4_title, news_4_link=news_4_link, news_5=news_5_title, news_5_link=news_5_link)

# Making the covid assiatance page
@app.route("/ai-covid-assistant", methods=['GET', 'POST'])
def covid_assistant():
    global msg, reply

    if request.method == "POST":
        webpage = request.form
        msg = str(webpage['msgInp'])
        # Important Variables
        vocab_size = 800
        embedding_dim = 14
        max_length = 30
        trunc_type = 'post'
        padding_type = 'post'
        oov_tok = "<OOV>"
        # Total size 29
        training_size = 29
        num_epochs = 110

        # Defining the class names
        class_names = ["What", "Variants", "Symptoms", "Feeling", "Precautions"]

        # Loading in the json data
        with open("chatbot-dataset.json", 'r') as f:
            datastore = json.load(f)

        # Making the seperate lists for the sentence and if it has profanity or not
        sentences = []
        labels = []

        # Looping over and grabbing all elements listed above
        for item in datastore:
            sentences.append(item['sentence'])
            labels.append(item['tag'])

        # Creating the training and the testing data
        training_sentences = sentences[0:training_size]
        testing_sentences = sentences[training_size:]
        training_labels = labels[0:training_size]
        testing_labels = labels[training_size:]
        print(testing_labels)
        print(training_labels)

        print(training_sentences)

        # Creating the keras tokenizer
        tokenizer = Tokenizer(num_words=vocab_size, oov_token=oov_tok)
        tokenizer.fit_on_texts(training_sentences)

        word_index = tokenizer.word_index

        # Converting the texts to sequencies using the tokenizer
        training_sequences = tokenizer.texts_to_sequences(training_sentences)
        training_padded = pad_sequences(
            training_sequences, maxlen=max_length, padding=padding_type, truncating=trunc_type)

        testing_sequences = tokenizer.texts_to_sequences(testing_sentences)
        testing_padded = pad_sequences(
            testing_sequences, maxlen=max_length, padding=padding_type, truncating=trunc_type)

        training_padded = np.array(training_padded)
        training_labels = np.array(training_labels)
        testing_padded = np.array(testing_padded)
        testing_labels = np.array(testing_labels)

        # Making the model
        model = tf.keras.Sequential([
            tf.keras.layers.Embedding(
                vocab_size, embedding_dim, input_length=max_length),
            tf.keras.layers.GlobalAveragePooling1D(),
            tf.keras.layers.Dense(1000, activation='relu'),
            tf.keras.layers.Dense(800, activation='relu'),
            tf.keras.layers.Dense(700, activation='relu'),
            tf.keras.layers.Dense(500, activation='relu'),
            tf.keras.layers.Dense(300, activation='relu'),
            tf.keras.layers.Dense(124, activation='relu'),
            tf.keras.layers.Dense(24, activation='relu'),
            tf.keras.layers.Dense(5, activation='sigmoid')
        ])

        # Loading the model
        model = keras.models.load_model("covid-assistance-ai/")

        # Compiling and fitting the model
        model.compile(loss='sparse_categorical_crossentropy',
                      optimizer='adam', metrics=['accuracy'])

        # Fitting the model
        # history = model.fit(training_padded, training_labels, epochs=num_epochs,
        #                     validation_data=(testing_padded, testing_labels), verbose=2)

        # Getting the model summary
        model.summary()

        # Saving the model
        model.save("covid-assistance-ai/")

        # Reversing the word index
        reverse_word_index = dict([(value, key)
                                   for (key, value) in word_index.items()])

        def decode_sentence(text):
            # Fancy Looking List Comprehension
            return ' '.join([reverse_word_index.get(i, '?') for i in text])

        sentence = []
        results = []
        sentence.append(msg)

        print(sentence)

        final_response = 0

        # Getting the predictions to display
        sequences = tokenizer.texts_to_sequences(sentence)
        padded = pad_sequences(sequences, maxlen=max_length,
                               padding=padding_type, truncating=trunc_type)
        # print(model.predict(padded))
        prediction = model.predict(padded)

        print(prediction)

        # Getting the prediction and returning it mapped to the clsss names list defined above
        response = class_names[np.argmax(prediction)]
        print(np.argmax(prediction))
        print(response)

        # Getting the response
        if response == "What":
            reply = "Coronavirus is a deadly virus that was originated in China in 2019."
            print("Coronavirus is a deadly virus that was originated in China in 2019.")

        elif response == "Variants":
            reply = "Alpha, Beta, Gamma and Delta are the main variants"
            print("Alpha, Beta, Gamma and Delta are the main variants")

        elif response == "Symptoms":
            reply = "Most common symptoms: fever dry cough tiredness Less common symptoms: aches and pains sore throat diarrhoea conjunctivitis headache loss of taste or smell a rash on skin, or discolouration of fingers or toes Serious symptoms: difficulty breathing or shortness of breath chest pain or pressure loss of speech or movement"
            print(
                "Most common symptoms: fever dry cough tiredness Less common symptoms: aches and pains sore throat diarrhoea conjunctivitis headache loss of taste or smell a rash on skin, or discolouration of fingers or toes Serious symptoms: difficulty breathing or shortness of breath chest pain or pressure loss of speech or movement")

        elif response == "Feeling":
            reply = "You should visit a doctor and check your oxygen levels and if they fall below 90 then you should seek immediate medical attention. Also check to see if the symptom that you are showing actually is a COVID symptom cause sometimes you will just have some common disease"
            print(
                "You should visit a doctor and check your oxygen levels and if they fall below 90 then you should seek immediate medical attention. Also check to see if the symptom that you are showing actually is a COVID symptom cause sometimes you will just have some common disease")

        elif response == "Precautions":
            reply = "You should always wear a mask and a face shield is optional"
            print("You should always wear a mask and a face shield is optional")

    return render_template("covid-assistance.html", msg=msg, reply=reply)

# Running the app
if __name__ == '__main__':
    app.run(debug=True)